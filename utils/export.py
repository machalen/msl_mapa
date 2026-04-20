from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="2563D4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for i, _ in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = max(
            (len(str(ws.cell(r, i).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


class ExcelExporter:
    def export_hospitals(self, data: list[dict], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hospitals"
        headers = ["Nom", "Status", "Contacte", "Observacions", "Lat", "Lng"]
        rows = [
            [d["nom"], d["status"], d.get("contacte", ""),
             d.get("observacions", ""), d["lat"], d["lng"]]
            for d in data
        ]
        _write_sheet(ws, headers, rows)
        wb.save(filepath)

    def export_doctors(self, data: list[dict], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Doctors"
        headers = ["Nom", "Especialitat", "Email", "Telèfon",
                   "Institució", "LinkedIn", "Observacions", "Hospital"]
        rows = [
            [d["nom"], d.get("especialitat", ""), d.get("email", ""),
             d.get("telefon", ""), d.get("institucio", ""), d.get("linkedin", ""),
             d.get("observacions", ""), d.get("hospital_nom", "")]
            for d in data
        ]
        _write_sheet(ws, headers, rows)
        wb.save(filepath)

    def export_projectes(self, data: list[dict], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projectes"
        headers = ["Nom", "Tema", "Status", "Observacions", "Hospital"]
        rows = [
            [d["nom"], d.get("tema", ""), d["status"],
             d.get("observacions", ""), d.get("hospital_nom", "")]
            for d in data
        ]
        _write_sheet(ws, headers, rows)
        wb.save(filepath)
